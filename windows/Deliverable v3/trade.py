# -*- coding: utf-8 -*-
"""
Created on Tue Nov 19 07:52:20 2019

@author: Nanda
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import xlwings as xw
import re
import time
import requests
import pandas as pd
import io
import os
import inspect
        
def get_data_alphavantage(symbol_list):
    # Update API key here
    api_key = 'VVFJ65QADWRFQDEQ'
    
    fun = 'TIME_SERIES_DAILY_ADJUSTED'
    nrow = len(symbol_list)
    
    print('Getting data for ' + str(nrow) + ' symbols')
    
    data_df = {}
    price_df = {}
    atr_df = {}
    error_tickers = []
    for i in range(nrow):
        symbol = symbol_list[i]
        print(str(i+1) + '. ' + symbol)
        
        try: 
            data_url = 'https://www.alphavantage.co/query?function=' + fun + '&symbol=' + symbol + '&outputsize=full' + '&apikey=' + api_key + '&datatype=csv'
            data = requests.get(data_url)
            data_df[symbol] = pd.read_csv(io.StringIO(data.content.decode('utf8')))
            price_df[symbol] = data_df[symbol]['adjusted_close'].iloc[:100]
    
            time.sleep(15)
            
            atr_period = 5
            atr_url = 'https://www.alphavantage.co/query?function=ATR&symbol=' + symbol + '&interval=daily&time_period=' + str(atr_period) + '&apikey=' + api_key + '&datatype=csv'
            atr_data = requests.get(atr_url)
            atr_df[symbol] = pd.read_csv(io.StringIO(atr_data.content.decode('utf8')))
            
            time.sleep(15)
            
        except:
            error_tickers = error_tickers.append(symbol)
            pass
    
    return price_df, data_df, atr_df

def green_red(data_df, lookback):
    data_df = data_df.iloc[:lookback,:]
    greens = data_df['close'] > data_df['open']
    reds = data_df['close'] < data_df['open']
    green_count = sum(greens)
    red_count = sum(reds)
    
    return green_count, red_count

def update_excel():
    # Set working directory    
    filename = inspect.getframeinfo(inspect.currentframe()).filename
    os.chdir(os.path.dirname(os.path.abspath(filename)))
    
    excel = 'OrdersFinal_new.xlsx'
    
    wb = load_workbook(excel)
    sheet = wb.active
    
    inputs = {}
    for row in sheet.iter_rows():
        for entry in row:
            try:
                if 'Order Size' in entry.value:
                    inputs['Order Size'] = entry.offset(row=1).value
                elif 'SL (xATR)' in entry.value:
                    inputs['SL'] = entry.offset(row=1).value
                elif 'TP' in entry.value:
                    inputs['TP'] = entry.offset(row=1).value
                elif 'Stop (xATR)' in entry.value:
                    inputs['STP'] = entry.offset(row=1).value
                elif 'Trail (xATR)' in entry.value:
                    inputs['Trail'] = entry.offset(row=1).value
                elif 'Lookback (days)' in entry.value:
                    inputs['Lookback'] = entry.offset(row=1).value
            
            except (AttributeError, TypeError):
                continue
            
    symbols = []
    signals = []
    for row in range(2, sheet.max_row+1):
        symbols.append(sheet['B' + str(row)].value)
        signals.append(sheet['A' + str(row)].value)    
    symbols = [x for x in symbols if x != None]
    signals = [x for x in signals if x != None]
    
    price_df, data_df, atr_df = get_data_alphavantage(symbols)
    
    price_list = []
    atr_list = []
    green_list = []
    red_list = []
    for i in range(len(symbols)):
        symbol = symbols[i]
        price_list.append(price_df[symbols[i]].iloc[0])
        atr_list.append(atr_df[symbols[i]].loc[:,'ATR'].iloc[0])
        greens, reds = green_red(data_df[symbol], inputs['Lookback'])
        green_list.append(greens)
        red_list.append(reds)
    
    for i in range(len(symbols)):
        sheet['C' + str(i+2)] = price_list[i]
        sheet['D' + str(i+2)] = atr_list[i]
        sheet['J' + str(i+2)] = green_list[i]
        sheet['K' + str(i+2)] = red_list[i]
    wb.save(excel)
    
    wb = xw.Book(excel)
    wb.save()
    wb.close()
    
    return symbols

def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

def get_order_pars():
    
    symbols = update_excel()
    
    # Set working directory    
    filename = inspect.getframeinfo(inspect.currentframe()).filename
    os.chdir(os.path.dirname(os.path.abspath(filename)))
    
    excel = 'OrdersFinal_new.xlsx'
    
    wb = load_workbook(excel, data_only = True)
    sheet = wb.active

    orders_df = load_workbook_range('A1:K' + str(len(symbols) + 1), sheet)
    orders_df.columns = orders_df.iloc[0,:]
    orders_df = orders_df.iloc[1:,:]
    
    for i in range(2, orders_df.shape[1]):
        orders_df.iloc[:,i] = orders_df.iloc[:,i].astype(float)
        orders_df.iloc[:,i] = round(orders_df.iloc[:,i], 2)
    orders_df['QTY'] = round(orders_df['QTY'])
        
    return orders_df

orders_df = get_order_pars()